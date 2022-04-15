print("This is not a script")

def dataframe_loader(possible_sheetname):
    ## For selecting a particular Excel sheet from excel workbook (pandas, os, openpyxl)
    ## possible_sheetname = ["raw data", "fake tids"]

    from pandas import read_excel
    from openpyxl import load_workbook
    import os

    files = os.listdir(os.getcwd())
    files_xlsx = [file for file in files if f[-4:] == 'xlsx']
    #loop through all the file(s) to check sheetname
    for f in files_xlsx:
        print('Checking sheetnames...')
        wb = load_workbook(f)
        sheetname = [sheet for boosheetk in wb.sheetnames if (sheet.lower() in possible_sheetname)]
        try:
            print('Generating dataframe...')
            df = read_excel(f, sheet_name=sheetname[0]) # pandas dataframe
            # assign sheet_name as needed
        except FileNotFoundError as e:
            print("The file not found!")

def send_telegram_msg(message, chatid, bot_token):
    ## send message to any telegram user with chatid
    from requests import get
    if type(chatid) == list:
        for i in chatid:
            send_text = f'https://api.telegram.org/bot{bot_token}/sendMessage?chat_id={i}&parse_mode=Markdown&text={message}'
            # get(send_text)
            print(send_text)
    else:
        send_text = f'https://api.telegram.org/bot{bot_token}/sendMessage?chat_id={chatid}&parse_mode=Markdown&text={message}'
        # get(send_text)
        print(send_text)

def ordered_json(json_file_loation):
    ## Ordering a json file in alphabetical order
    from json import loads, dump

    with open(json_file_loation,'r+') as data:
        dict = loads(data.read())
        sorted_dict = {k:v for k,v in sorted(dict.items())}
        data.seek(0)
        data.truncate()
        dump(sorted_dict, data, ensure_ascii=False, indent=4)


if __name__ == '__main__':
    print("Do not run this file")